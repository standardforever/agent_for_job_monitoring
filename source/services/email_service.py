from __future__ import annotations

import asyncio
import base64
import csv
import hashlib
import html
import io
import json
import mimetypes
import os
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from textwrap import dedent
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.config import get_settings
from utils.logging import get_logger, log_event

logger = get_logger("email_service")

API_URL = "https://api.resend.com/emails"
ALLOWED_EXTENSIONS = {".pdf", ".csv", ".xlsx"}


@dataclass(slots=True)
class EmailAttachment:
    filename: str
    content: bytes
    content_type: str = "application/octet-stream"


def _json_default(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


IMPORTANT_CSV_FIELDS = [
    "client_name",
    "status",
    "raw_url",
    "result_summary_career_url_status",
    "career_page_overview_outcome",
    "career_page_overview_outcome_reason",
    "career_page_overview_job_found_on_urls",
    "extracted_job_count",
    "new_job_count",
]

ROLES_CSV_FIELDS = [
    "company_url",
    "domain_key",
    "career_url",
    "listing_page_url",
    "job_url",
    "title",
    "source",
    "page_status",
    "snapshot_run_date",
    "snapshot_job_count",
    "snapshot_added_count",
    "snapshot_removed_count",
    "snapshot_unchanged_count",
    "job_change_status",
]

ALERT_JOBS_CSV_FIELDS = [
    "Company",
    "Job Title",
    "Date Detected",
    "Apply URL",
    "Source URL",
]

C_HEADER_BG = "0D1B40"
C_SECTION_BG = "1249A0"
C_COL_HDR_BG = "1565C0"
C_DATA_BG = "EBF2FF"
C_DATA_SIDE_BG = "F4F7FF"
C_FOOTER_BG = "EEF3FB"
C_WHITE = "FFFFFF"
C_SUBTITLE = "8BAED4"
C_PERIOD = "6A9FD8"
C_DARK_TEXT = "1A1A2E"
C_DATE_TEXT = "555577"
C_LINK = "1565C0"
C_FOOTER_TEXT = "8899BB"
C_BORDER_GRAY = "AAAAAA"
C_BORDER_BLUE = "0D47A1"
C_BORDER_LIGHT = "C5D3F0"


def _csv_content(rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> str:
    output = io.StringIO()
    resolved_fieldnames = fieldnames or (list(rows[0].keys()) if rows else [])
    writer = csv.DictWriter(output, fieldnames=resolved_fieldnames, extrasaction="ignore")
    if resolved_fieldnames:
        writer.writeheader()
        for row in rows:
            writer.writerow({key: _json_default(value) for key, value in row.items()})
    content = output.getvalue()
    output.close()
    return content


def _safe_filename(value: str) -> str:
    return "".join(character if character.isalnum() or character in {"-", "_", "."} else "_" for character in value)


def _flatten_csv_list(values: Any) -> str:
    flattened: list[str] = []
    for item in list(values or []):
        if isinstance(item, (dict, list)):
            flattened.append(json.dumps(item, ensure_ascii=False, default=_json_default))
        elif item is not None:
            flattened.append(str(item))
    return " | ".join(value for value in flattened if value)


def _format_report_datetime(value: Any) -> str:
    raw_value = str(value or "").strip()
    if not raw_value:
        return ""
    try:
        parsed = datetime.fromisoformat(raw_value.replace("Z", "+00:00"))
    except ValueError:
        return raw_value
    return parsed.strftime("%d %b %Y  %H:%M")


def _report_side(style: str | None, color: str | None = None) -> Side:
    return Side(border_style=style, color=color)


NO_REPORT_SIDE = _report_side(None)
THIN_REPORT_GRAY = _report_side("thin", C_BORDER_GRAY)
MED_REPORT_BLUE = _report_side("medium", C_BORDER_BLUE)
HAIR_REPORT_LIGHT = _report_side("hair", C_BORDER_LIGHT)


def _report_border(
    *,
    left: Side = NO_REPORT_SIDE,
    right: Side = NO_REPORT_SIDE,
    top: Side = NO_REPORT_SIDE,
    bottom: Side = NO_REPORT_SIDE,
) -> Border:
    return Border(left=left, right=right, top=top, bottom=bottom)


def _report_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _report_font(
    *,
    name: str = "Arial",
    size: int = 10,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    color: str = C_DARK_TEXT,
) -> Font:
    return Font(
        name=name,
        size=size,
        bold=bold,
        italic=italic,
        underline="single" if underline else None,
        color=color,
    )


def _report_align(horizontal: str = "left", vertical: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def _set_report_cell(
    worksheet: Any,
    coordinate: str,
    value: Any = None,
    *,
    fnt: Font | None = None,
    fll: PatternFill | None = None,
    aln: Alignment | None = None,
    brd: Border | None = None,
) -> None:
    cell = worksheet[coordinate]
    if value is not None:
        cell.value = value
    if fnt:
        cell.font = fnt
    if fll:
        cell.fill = fll
    if aln:
        cell.alignment = aln
    if brd:
        cell.border = brd


def _paint_report_outer_row(
    worksheet: Any,
    row: int,
    fill_color: str | None,
    *,
    top: Side = NO_REPORT_SIDE,
    bottom: Side = NO_REPORT_SIDE,
) -> None:
    row_fill = _report_fill(fill_color) if fill_color else None
    for column in "BCDE":
        _set_report_cell(worksheet, f"{column}{row}", fll=row_fill, brd=_report_border(top=top, bottom=bottom))
    _set_report_cell(worksheet, f"A{row}", fll=row_fill, brd=_report_border(left=THIN_REPORT_GRAY, top=top, bottom=bottom))
    _set_report_cell(worksheet, f"F{row}", fll=row_fill, brd=_report_border(right=THIN_REPORT_GRAY, top=top, bottom=bottom))


def _format_alert_period(value: Any) -> str:
    formatted = _format_report_datetime(value)
    return formatted.replace("  ", ", ", 1) if formatted else ""


def _alert_report_jobs(alert: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for job in list(alert.get("jobs_after_filter") or []):
        rows.append(
            {
                "company": str(job.get("registered_domain") or "").strip(),
                "job_title": str(job.get("title") or "").strip(),
                "date_detected": _format_report_datetime(job.get("first_seen_at")),
                "apply_url": str(job.get("job_url") or "").strip(),
            }
        )
    return rows


def _build_alert_jobs_xlsx(alert: dict[str, Any]) -> bytes:
    jobs = _alert_report_jobs(alert)
    generated_date = _format_report_datetime(alert.get("period_end") or datetime.now()).split("  ")[0]
    report_start = _format_alert_period(alert.get("period_start")) or generated_date
    report_end = _format_alert_period(alert.get("period_end")) or generated_date
    branding = "ClickBuy.ai"
    vacancy_word = "vacancy" if len(jobs) == 1 else "vacancies"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "New Vacancies"
    worksheet.sheet_view.showGridLines = False

    worksheet.column_dimensions["A"].width = 3
    worksheet.column_dimensions["B"].width = 28
    worksheet.column_dimensions["C"].width = 48
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 22
    worksheet.column_dimensions["F"].width = 3

    worksheet.row_dimensions[1].height = 8
    _paint_report_outer_row(worksheet, 1, C_HEADER_BG, top=THIN_REPORT_GRAY)

    worksheet.row_dimensions[2].height = 38
    _paint_report_outer_row(worksheet, 2, C_HEADER_BG)
    worksheet.merge_cells("B2:E2")
    _set_report_cell(
        worksheet,
        "B2",
        value="🔍  HIRING SIGNAL REPORT",
        fnt=_report_font(size=20, bold=True, color=C_WHITE),
        fll=_report_fill(C_HEADER_BG),
        aln=_report_align("left", "center"),
    )

    worksheet.row_dimensions[3].height = 20
    _paint_report_outer_row(worksheet, 3, C_HEADER_BG)
    worksheet.merge_cells("B3:E3")
    _set_report_cell(
        worksheet,
        "B3",
        value=f"New vacancies detected in the last 24 hours · Powered by {branding}",
        fnt=_report_font(size=10, color=C_SUBTITLE),
        fll=_report_fill(C_HEADER_BG),
        aln=_report_align("left", "center"),
    )

    worksheet.row_dimensions[4].height = 20
    _paint_report_outer_row(worksheet, 4, C_HEADER_BG)
    worksheet.merge_cells("B4:E4")
    _set_report_cell(
        worksheet,
        "B4",
        value=f"Report period:  {report_start}  →  {report_end}",
        fnt=_report_font(size=9, color=C_PERIOD),
        fll=_report_fill(C_HEADER_BG),
        aln=_report_align("left", "center"),
    )

    worksheet.row_dimensions[5].height = 10
    _paint_report_outer_row(worksheet, 5, C_HEADER_BG)

    worksheet.row_dimensions[6].height = 30
    _set_report_cell(worksheet, "A6", fll=_report_fill(C_SECTION_BG), brd=_report_border(left=THIN_REPORT_GRAY))
    worksheet.merge_cells("B6:C6")
    _set_report_cell(
        worksheet,
        "B6",
        value=f"📋  {len(jobs)} new {vacancy_word} found",
        fnt=_report_font(size=11, bold=True, color=C_WHITE),
        fll=_report_fill(C_SECTION_BG),
        aln=_report_align("left", "center"),
    )
    _set_report_cell(worksheet, "C6", fll=_report_fill(C_SECTION_BG))
    worksheet.merge_cells("D6:E6")
    _set_report_cell(
        worksheet,
        "D6",
        value=f"Generated: {generated_date}",
        fnt=_report_font(size=9, color=C_SUBTITLE),
        fll=_report_fill(C_SECTION_BG),
        aln=_report_align("right", "center"),
    )
    _set_report_cell(worksheet, "E6", fll=_report_fill(C_SECTION_BG))
    _set_report_cell(worksheet, "F6", fll=_report_fill(C_SECTION_BG), brd=_report_border(right=THIN_REPORT_GRAY))

    worksheet.row_dimensions[7].height = 8
    _set_report_cell(worksheet, "A7", brd=_report_border(left=THIN_REPORT_GRAY))
    _set_report_cell(worksheet, "F7", brd=_report_border(right=THIN_REPORT_GRAY))

    worksheet.row_dimensions[8].height = 28
    header_font = _report_font(size=10, bold=True, color=C_DARK_TEXT)
    header_fill = _report_fill(C_COL_HDR_BG)
    header_align = _report_align("left", "center")
    header_border = _report_border(bottom=MED_REPORT_BLUE)
    _set_report_cell(worksheet, "A8", fll=header_fill, brd=_report_border(left=THIN_REPORT_GRAY))
    for coordinate, label in [("B8", "Company"), ("C8", "Job Title"), ("D8", "Date Detected"), ("E8", "Apply")]:
        _set_report_cell(
            worksheet,
            coordinate,
            value=label,
            fnt=header_font,
            fll=header_fill,
            aln=header_align,
            brd=header_border,
        )
    _set_report_cell(worksheet, "F8", fll=header_fill, brd=_report_border(right=THIN_REPORT_GRAY))

    for index, job in enumerate(jobs):
        row = 9 + index
        worksheet.row_dimensions[row].height = 24
        row_border = _report_border(top=MED_REPORT_BLUE, bottom=HAIR_REPORT_LIGHT)
        _set_report_cell(worksheet, f"A{row}", fll=_report_fill(C_DATA_SIDE_BG), brd=_report_border(left=THIN_REPORT_GRAY))
        _set_report_cell(
            worksheet,
            f"B{row}",
            value=job["company"],
            fnt=_report_font(size=10, bold=True, color=C_DARK_TEXT),
            fll=_report_fill(C_DATA_BG),
            aln=_report_align("left", "center"),
            brd=row_border,
        )
        _set_report_cell(
            worksheet,
            f"C{row}",
            value=job["job_title"],
            fnt=_report_font(size=10, color=C_DARK_TEXT),
            fll=_report_fill(C_DATA_BG),
            aln=_report_align("left", "center"),
            brd=row_border,
        )
        _set_report_cell(
            worksheet,
            f"D{row}",
            value=job["date_detected"],
            fnt=_report_font(size=9, color=C_DATE_TEXT),
            fll=_report_fill(C_DATA_BG),
            aln=_report_align("left", "center"),
            brd=row_border,
        )
        apply_cell = worksheet[f"E{row}"]
        apply_cell.value = "View vacancy →"
        apply_cell.font = _report_font(size=9, underline=True, color=C_LINK)
        apply_cell.fill = _report_fill(C_DATA_BG)
        apply_cell.alignment = _report_align("left", "center")
        apply_cell.border = row_border
        if job["apply_url"]:
            apply_cell.hyperlink = job["apply_url"]
        _set_report_cell(worksheet, f"F{row}", fll=_report_fill(C_DATA_SIDE_BG), brd=_report_border(right=THIN_REPORT_GRAY))

    post_row = 9 + len(jobs)
    worksheet.row_dimensions[post_row].height = 13.55
    _set_report_cell(worksheet, f"A{post_row}", brd=_report_border(left=THIN_REPORT_GRAY))
    for column in "BCDE":
        _set_report_cell(worksheet, f"{column}{post_row}", brd=_report_border(top=HAIR_REPORT_LIGHT))
    _set_report_cell(worksheet, f"F{post_row}", brd=_report_border(right=THIN_REPORT_GRAY))

    spacer_row = post_row + 1
    worksheet.row_dimensions[spacer_row].height = 10
    _set_report_cell(worksheet, f"A{spacer_row}", brd=_report_border(left=THIN_REPORT_GRAY))
    _set_report_cell(worksheet, f"F{spacer_row}", brd=_report_border(right=THIN_REPORT_GRAY))

    footer_row = spacer_row + 1
    worksheet.row_dimensions[footer_row].height = 28
    footer_text = (
        f"This report was automatically generated by the {branding} Hiring Signal Agent"
        "  ·  Reply to find out how this could work for your desk"
    )
    _set_report_cell(worksheet, f"A{footer_row}", fll=_report_fill(C_FOOTER_BG), brd=_report_border(left=THIN_REPORT_GRAY, bottom=THIN_REPORT_GRAY))
    worksheet.merge_cells(f"B{footer_row}:E{footer_row}")
    _set_report_cell(
        worksheet,
        f"B{footer_row}",
        value=footer_text,
        fnt=_report_font(size=8, italic=True, color=C_FOOTER_TEXT),
        fll=_report_fill(C_FOOTER_BG),
        aln=_report_align("center", "center"),
        brd=_report_border(bottom=THIN_REPORT_GRAY),
    )
    for column in "CDE":
        _set_report_cell(worksheet, f"{column}{footer_row}", fll=_report_fill(C_FOOTER_BG), brd=_report_border(bottom=THIN_REPORT_GRAY))
    _set_report_cell(worksheet, f"F{footer_row}", fll=_report_fill(C_FOOTER_BG), brd=_report_border(right=THIN_REPORT_GRAY, bottom=THIN_REPORT_GRAY))

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_career_outcome_reason_with_pagination(career_overview: dict[str, Any]) -> str | None:
    reason = str(career_overview.get("outcome_reason") or "").strip() or None
    listing_ui = career_overview.get("listing_ui")
    pagination_present = listing_ui.get("pagination_present") if isinstance(listing_ui, dict) else None
    if pagination_present is True:
        suffix = " Pagination detected on the page: yes."
    elif pagination_present is False:
        suffix = " Pagination detected on the page: no."
    else:
        suffix = " Pagination detected on the page: unknown."
    return f"{reason or ''}{suffix}".strip() or None


def build_process_important_csv_rows(process: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in list(process.get("items") or []):
        result_payload = dict(item.get("result_payload") or {})
        career_page_result = dict(result_payload.get("career_page_result") or {})
        result_summary = dict(item.get("result_summary") or {})
        career_overview = dict(career_page_result.get("overview") or {})
        rows.append(
            {
                "client_name": process.get("client_name"),
                "status": item.get("status"),
                "raw_url": item.get("raw_url"),
                "result_summary_career_url_status": result_summary.get("career_url_status"),
                "career_page_overview_outcome": career_overview.get("outcome"),
                "career_page_overview_outcome_reason": _build_career_outcome_reason_with_pagination(career_overview),
                "career_page_overview_job_found_on_urls": _flatten_csv_list(career_overview.get("job_found_on_urls")),
                "extracted_job_count": result_summary.get("extracted_job_count"),
                "new_job_count": result_summary.get("new_job_count"),
            }
        )
    return rows


def _listing_page_url(page: dict[str, Any]) -> str:
    pattern_result = page.get("job_listing_pattern")
    pattern_page_url = pattern_result.get("page_url") if isinstance(pattern_result, dict) else None
    return str(
        page.get("job_listing_pattern_url")
        or page.get("classified_job_listing_url")
        or pattern_page_url
        or page.get("extracted_url")
        or page.get("current_url")
        or page.get("url")
        or page.get("navigation_url")
        or ""
    ).strip()


def _career_url(page: dict[str, Any]) -> str:
    return str(
        page.get("extracted_url")
        or page.get("current_url")
        or page.get("url")
        or page.get("navigation_url")
        or _listing_page_url(page)
        or ""
    ).strip()


def _job_title(job: Any) -> str | None:
    if isinstance(job, dict):
        return str(job.get("title") or job.get("job_title") or "").strip() or None
    return None


def _job_url(job: Any) -> str | None:
    if isinstance(job, dict):
        return str(job.get("job_url") or job.get("url") or "").strip() or None
    return str(job or "").strip() or None


def _listing_job_key(domain_key: str | None, page_url: str | None, job: Any) -> str:
    title = _job_title(job) or ""
    job_url = _job_url(job) or ""
    key_source = job_url or f"{page_url or ''}|{title}".lower()
    source = f"{domain_key or ''}|{key_source}".strip().lower()
    return hashlib.sha256(source.encode("utf-8")).hexdigest()


def _process_is_rerun(process: dict[str, Any]) -> bool:
    metadata = dict(process.get("metadata") or {})
    if metadata.get("rerun_of_process_id") or metadata.get("workflow_mode") == "rerun":
        return True
    if process.get("history"):
        return True
    return any(item.get("previous_job_keys") for item in list(process.get("items") or []))


def _append_role_row(
    rows: list[dict[str, Any]],
    seen: set[tuple[str, str, str, str, str]],
    *,
    process: dict[str, Any],
    item: dict[str, Any],
    page: dict[str, Any],
    job: Any,
    source: str,
    new_only: bool,
) -> None:
    title = _job_title(job)
    job_url = _job_url(job)
    if not title and not job_url:
        return

    snapshot = dict(page.get("listing_job_snapshot") or {})
    listing_page_url = _listing_page_url(page) or None
    job_key = _listing_job_key(item.get("domain_key"), listing_page_url, job)
    added_job_keys = set(item.get("added_job_keys") or [])
    if new_only and job_key not in added_job_keys:
        return

    row = {
        "company_url": item.get("raw_url"),
        "domain_key": item.get("domain_key"),
        "career_url": _career_url(page) or None,
        "listing_page_url": listing_page_url,
        "job_url": job_url,
        "title": title,
        "source": source,
        "page_status": page.get("status"),
        "snapshot_run_date": snapshot.get("run_date"),
        "snapshot_job_count": snapshot.get("job_count"),
        "snapshot_added_count": len(snapshot.get("added_job_keys") or []),
        "snapshot_removed_count": len(snapshot.get("removed_job_keys") or []),
        "snapshot_unchanged_count": len(snapshot.get("unchanged_job_keys") or []),
        "job_change_status": "new" if new_only else "current",
    }
    if job_url:
        marker = (
            str(row["company_url"] or ""),
            "",
            "",
            str(job_url),
            "",
        )
    else:
        marker = (
            str(row["company_url"] or ""),
            str(row["career_url"] or ""),
            str(row["listing_page_url"] or ""),
            "",
            str(row["title"] or ""),
        )
    if marker in seen:
        return
    seen.add(marker)
    rows.append(row)


def _append_extracted_job_row(
    rows: list[dict[str, Any]],
    seen: set[tuple[str, str, str, str, str]],
    *,
    item: dict[str, Any],
    job: dict[str, Any],
    new_only: bool,
) -> None:
    job_key = str(job.get("job_key") or "").strip()
    if new_only and job_key not in set(item.get("added_job_keys") or []):
        return

    title = _job_title(job)
    job_url = _job_url(job)
    if not title and not job_url:
        return

    marker = (
        str(item.get("raw_url") or ""),
        str(item.get("career_page_url") or ""),
        str(job.get("listing_page_url") or ""),
        str(job_url or ""),
        str(title or ""),
    )
    if marker in seen:
        return
    seen.add(marker)
    rows.append(
        {
            "company_url": item.get("raw_url"),
            "domain_key": item.get("domain_key"),
            "career_url": item.get("career_page_url"),
            "listing_page_url": job.get("listing_page_url"),
            "job_url": job_url,
            "title": title,
            "source": job.get("source") or "extracted_jobs",
            "page_status": item.get("status"),
            "snapshot_run_date": (item.get("completed_at") or item.get("updated_at")),
            "snapshot_job_count": len(item.get("current_job_keys") or []),
            "snapshot_added_count": len(item.get("added_job_keys") or []),
            "snapshot_removed_count": len(item.get("removed_job_keys") or []),
            "snapshot_unchanged_count": len(item.get("unchanged_job_keys") or []),
            "job_change_status": "new" if new_only else "current",
        }
    )


def build_process_roles_csv_rows(process: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str, str]] = set()
    new_only = _process_is_rerun(process)
    for item in list(process.get("items") or []):
        result_payload = dict(item.get("result_payload") or {})
        extracted_jobs = list(result_payload.get("extracted_jobs") or [])
        if extracted_jobs:
            for job in extracted_jobs:
                if isinstance(job, dict):
                    _append_extracted_job_row(rows, seen, item=item, job=job, new_only=new_only)
            continue
        if new_only and not item.get("added_job_keys"):
            continue

        career_page_result = dict(result_payload.get("career_page_result") or {})
        item_rows_before = len(rows)
        for page in list(career_page_result.get("career_pages_analysis") or []):
            page_rows_before = len(rows)
            llm_analysis = dict(page.get("llm_analysis") or {})
            for job in list(llm_analysis.get("jobs_listed_on_page") or []):
                _append_role_row(
                    rows,
                    seen,
                    process=process,
                    item=item,
                    page=page,
                    job=job,
                    source="llm_analysis",
                    new_only=new_only,
                )

            pattern_result = page.get("job_listing_pattern")
            if isinstance(pattern_result, dict):
                for job in list(pattern_result.get("jobs") or []):
                    _append_role_row(
                        rows,
                        seen,
                        process=process,
                        item=item,
                        page=page,
                        job=job,
                        source="job_listing_pattern",
                        new_only=new_only,
                    )

            if len(rows) == page_rows_before:
                for job_url in list(page.get("jobs_listed_on_page") or []):
                    _append_role_row(
                        rows,
                        seen,
                        process=process,
                        item=item,
                        page=page,
                        job=job_url,
                        source="page_jobs_listed",
                        new_only=new_only,
                    )

        overview = dict(career_page_result.get("overview") or {})
        if not new_only and len(rows) == item_rows_before:
            overview_page = {
                "status": "overview",
                "extracted_url": None,
                "classified_job_listing_url": (overview.get("job_found_on_urls") or [None])[0],
            }
            for job_url in list(overview.get("job_urls") or []):
                _append_role_row(
                    rows,
                    seen,
                    process=process,
                    item=item,
                    page=overview_page,
                    job=job_url,
                    source="career_page_overview",
                    new_only=new_only,
                )
    return rows


def build_process_csv_attachments(process: dict[str, Any]) -> list[EmailAttachment]:
    process_id = str(process.get("process_id") or "process").strip() or "process"
    safe_process_id = _safe_filename(process_id)
    return [
        EmailAttachment(
            filename=f"process_{safe_process_id}_important.csv",
            content=_csv_content(build_process_important_csv_rows(process), IMPORTANT_CSV_FIELDS).encode(),
            content_type="text/csv",
        ),
        EmailAttachment(
            filename=f"process_{safe_process_id}_roles.csv",
            content=_csv_content(build_process_roles_csv_rows(process), ROLES_CSV_FIELDS).encode(),
            content_type="text/csv",
        ),
    ]


def build_alert_jobs_csv_attachment(alert: dict[str, Any]) -> EmailAttachment:
    return build_alert_jobs_report_attachment(alert)


def build_alert_jobs_report_attachment(alert: dict[str, Any]) -> EmailAttachment:
    client = dict(alert.get("client") or {})
    period_type = str(alert.get("period_type") or "daily").strip() or "daily"
    client_name = _safe_filename(str(client.get("client_name") or "client"))
    alert_id = _safe_filename(str(alert.get("alert_id") or "alert"))
    return EmailAttachment(
        filename=f"{period_type}_hiring_signal_report_{client_name}_{alert_id}.xlsx",
        content=_build_alert_jobs_xlsx(alert),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def build_html_email(body: str) -> str:
    paragraphs = [
        f"<p>{html.escape(block).replace(chr(10), '<br>')}</p>"
        for block in body.split("\n\n")
        if block.strip()
    ]
    body_html = "\n".join(paragraphs)
    return dedent(
        f"""
        <!doctype html>
        <html lang="en">
          <body style="margin:0;padding:0;background:#f5f7fa;font-family:Arial,Helvetica,sans-serif;color:#172033;">
            <div style="display:none;max-height:0;overflow:hidden;color:transparent;">
              Your ProcessZero career page job report is ready.
            </div>

            <div style="max-width:720px;margin:0 auto;padding:32px 18px;">
              <div style="padding:26px 30px;border-radius:14px 14px 0 0;background:#111827;color:#ffffff;">
                <div style="font-size:13px;line-height:1.4;letter-spacing:0;text-transform:uppercase;color:#b9c3d3;">
                  ProcessZero
                </div>
                <h1 style="margin:8px 0 0;font-size:26px;line-height:1.25;font-weight:700;">
                  Career page job report
                </h1>
              </div>

              <div style="padding:30px;border:1px solid #e4e9f0;border-top:0;background:#ffffff;font-size:15px;line-height:1.75;">
                {body_html}

                <div style="margin-top:28px;padding-top:22px;border-top:1px solid #e4e9f0;">
                  <a href="https://processzero.co.uk/" style="display:inline-block;padding:11px 16px;border-radius:8px;background:#111827;color:#ffffff;text-decoration:none;font-size:14px;font-weight:700;">
                    Visit ProcessZero
                  </a>
                </div>
              </div>

              <div style="padding:18px 30px;border:1px solid #e4e9f0;border-top:0;border-radius:0 0 14px 14px;background:#fbfcfe;color:#697386;font-size:12px;line-height:1.6;">
                Sent by ProcessZero<br>
                <a href="https://processzero.co.uk/" style="color:#334155;text-decoration:none;">https://processzero.co.uk/</a>
              </div>
            </div>
          </body>
        </html>
        """
    ).strip()


def _attachment_from_path(path_str: str) -> EmailAttachment:
    path = Path(path_str)
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"Attachment not found: {path}")
    if path.suffix.lower() not in ALLOWED_EXTENSIONS:
        allowed = ", ".join(sorted(ALLOWED_EXTENSIONS))
        raise ValueError(f"Unsupported attachment type for {path.name}. Allowed: {allowed}")
    mime_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    return EmailAttachment(filename=path.name, content=path.read_bytes(), content_type=mime_type)


def _serialize_attachment(attachment: EmailAttachment | str) -> dict[str, str]:
    normalized = _attachment_from_path(attachment) if isinstance(attachment, str) else attachment
    return {
        "filename": normalized.filename,
        "content": base64.b64encode(normalized.content).decode("utf-8"),
        "content_type": normalized.content_type,
    }


def format_from_address(from_email: str, from_name: str | None = None) -> str:
    if not from_name:
        return from_email.strip()
    safe_name = " ".join(from_name.strip().split())
    return f"{safe_name} <{from_email.strip()}>"


def send_email(
    *,
    from_email: str,
    to: list[str],
    subject: str,
    body: str,
    from_name: str | None = None,
    reply_to: str | None = None,
    attachments: list[EmailAttachment | str] | None = None,
    api_key: str | None = None,
) -> dict[str, Any]:
    api_key = api_key or os.getenv("RESEND_API_KEY")
    clean_to = [email.strip() for email in to if email and email.strip()]
    if not api_key:
        raise RuntimeError("RESEND_API_KEY is not set.")
    if not from_email or not from_email.strip():
        raise RuntimeError("Email from address is not configured.")
    if not clean_to:
        raise RuntimeError("No email recipients provided.")

    payload: dict[str, Any] = {
        "from": format_from_address(from_email, from_name),
        "to": clean_to,
        "subject": subject,
        "html": build_html_email(body),
        "text": body,
    }
    if reply_to:
        payload["reply_to"] = reply_to
    if attachments:
        payload["attachments"] = [_serialize_attachment(path) for path in attachments]

    log_event(
        logger,
        "info",
        "email_send_started recipient_count=%s attachment_count=%s subject=%s",
        len(clean_to),
        len(attachments or []),
        subject,
        domain="email",
        recipient_count=len(clean_to),
        attachment_count=len(attachments or []),
        subject=subject,
    )
    response = requests.post(
        API_URL,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json=payload,
        timeout=30,
    )
    response.raise_for_status()
    result = response.json()
    log_event(
        logger,
        "info",
        "email_send_completed resend_id=%s",
        result.get("id"),
        domain="email",
        resend_id=result.get("id"),
        recipient_count=len(clean_to),
    )
    return result


class EmailService:
    def __init__(self) -> None:
        self._settings = get_settings()

    async def send_process_completed_email(
        self,
        *,
        client: dict[str, Any],
        process: dict[str, Any],
    ) -> dict[str, Any]:
        return await asyncio.to_thread(self._send_process_completed_email_sync, client, process)

    def _send_process_completed_email_sync(
        self,
        client: dict[str, Any],
        process: dict[str, Any],
    ) -> dict[str, Any]:
        process_id = str(process.get("process_id") or "").strip()
        client_email = str(client.get("email") or "").strip()
        status = str(process.get("status") or "").strip()

        if not self._settings.process_email_enabled:
            log_event(
                logger,
                "info",
                "process_email_skipped_disabled process_id=%s",
                process_id,
                domain=process_id or "email",
                process_id=process_id,
            )
            return {"status": "skipped", "reason": "email_disabled"}
        if status not in {"completed", "partial_completed"}:
            log_event(
                logger,
                "info",
                "process_email_skipped_status process_id=%s status=%s",
                process_id,
                status,
                domain=process_id or "email",
                process_id=process_id,
                status=status,
            )
            return {"status": "skipped", "reason": "process_not_completed"}
        if not client_email:
            log_event(
                logger,
                "warning",
                "process_email_skipped_no_client_email process_id=%s client_key=%s",
                process_id,
                client.get("client_key"),
                domain=process_id or "email",
                process_id=process_id,
                client_key=client.get("client_key"),
            )
            return {"status": "skipped", "reason": "missing_client_email"}

        summary = dict(process.get("summary") or {})
        subject_prefix = str(self._settings.process_email_subject_prefix or "").strip()
        status_label = "partially completed" if status == "partial_completed" else "completed"
        subject = f"{subject_prefix} Process {process_id} {status_label}".strip()
        body = (
            f"Hello {client.get('client_name') or 'there'},\n\n"
            f"Your career page job extraction process has {status_label}.\n\n"
            f"Process ID: {process_id}\n"
            f"Status: {status}\n"
            f"Total domains: {summary.get('total_domain_count')}\n"
            f"Jobs found: {summary.get('job_count')}\n"
            f"New jobs: {summary.get('new_job_count')}\n"
            f"Completed domains: {summary.get('completed_domain_count')}\n"
            f"Failed domains: {summary.get('failed_domain_count')}\n\n"
            "Attached are two CSV files: the process summary and the roles found."
        )

        try:
            attachments = build_process_csv_attachments(process)
            result = send_email(
                from_email=self._settings.email_from_address,
                from_name=self._settings.email_from_name,
                reply_to=self._settings.email_reply_to or None,
                to=[client_email],
                subject=subject,
                body=body,
                attachments=attachments,
                api_key=self._settings.resend_api_key,
            )
            return {"status": "sent", "provider_response": result}
        except Exception as exc:
            log_event(
                logger,
                "warning",
                "process_email_failed process_id=%s client_key=%s error=%s",
                process_id,
                client.get("client_key"),
                str(exc),
                domain=process_id or "email",
                process_id=process_id,
                client_key=client.get("client_key"),
                error=str(exc),
            )
            return {"status": "failed", "error": str(exc)}
