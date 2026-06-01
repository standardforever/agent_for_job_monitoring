from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from utils.logging import get_logger, log_event

logger = get_logger("file_input_service")


@dataclass(slots=True)
class UploadDomainInput:
    domain: str
    career_url: str | None = None


class FileInputService:
    def extract_domain_inputs(self, filename: str, content: bytes) -> list[UploadDomainInput]:
        suffix = Path(filename).suffix.lower()
        log_event(
            logger,
            "info",
            "file_domain_input_extraction_started filename=%s suffix=%s",
            filename,
            suffix,
            domain=filename,
            upload_filename=filename,
            suffix=suffix,
        )

        if suffix == ".csv":
            inputs = self._extract_domain_inputs_from_csv(content)
        elif suffix == ".xlsx":
            inputs = self._extract_domain_inputs_from_xlsx(content)
        else:
            raise ValueError("Only .csv and .xlsx files are supported")

        if not inputs:
            raise ValueError("No valid values found in the 'domain' column")

        supplied_career_url_count = sum(1 for item in inputs if item.career_url)
        log_event(
            logger,
            "info",
            "file_domain_input_extraction_completed filename=%s domain_count=%s supplied_career_url_count=%s",
            filename,
            len(inputs),
            supplied_career_url_count,
            domain=inputs[0].domain,
            upload_filename=filename,
            domain_count=len(inputs),
            supplied_career_url_count=supplied_career_url_count,
        )
        return inputs

    def extract_domains(self, filename: str, content: bytes) -> list[str]:
        return [item.domain for item in self.extract_domain_inputs(filename, content)]

    def _extract_domain_inputs_from_csv(self, content: bytes) -> list[UploadDomainInput]:
        text_stream = StringIO(content.decode("utf-8-sig"))
        reader = csv.DictReader(text_stream)
        return self._collect_domain_inputs(reader)

    def _extract_domain_inputs_from_xlsx(self, content: bytes) -> list[UploadDomainInput]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not rows:
            return []

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        normalized_headers = [header.lower() for header in headers]
        if "domain" not in normalized_headers:
            raise ValueError("Uploaded file must contain a 'domain' column")

        domain_index = normalized_headers.index("domain")
        career_url_index = self._optional_header_index(normalized_headers, ["career_url", "career_page_url"])
        inputs: list[UploadDomainInput] = []
        seen: set[str] = set()
        for row in rows[1:]:
            if row is None or domain_index >= len(row):
                continue
            domain = str(row[domain_index] or "").strip()
            career_url = (
                str(row[career_url_index] or "").strip()
                if career_url_index is not None and career_url_index < len(row)
                else ""
            )
            marker = f"{domain}|{career_url}"
            if not domain or marker in seen:
                continue
            seen.add(marker)
            inputs.append(UploadDomainInput(domain=domain, career_url=career_url or None))
        return inputs

    def _collect_domain_inputs(self, reader: csv.DictReader) -> list[UploadDomainInput]:
        if reader.fieldnames is None:
            return []

        normalized_fieldnames = {str(name).strip().lower(): name for name in reader.fieldnames if name}
        if "domain" not in normalized_fieldnames:
            raise ValueError("Uploaded file must contain a 'domain' column")

        domain_key = normalized_fieldnames["domain"]
        career_url_key = self._optional_header_name(normalized_fieldnames, ["career_url", "career_page_url"])
        inputs: list[UploadDomainInput] = []
        seen: set[str] = set()
        for row in reader:
            domain = str((row or {}).get(domain_key) or "").strip()
            career_url = str((row or {}).get(career_url_key) or "").strip() if career_url_key else ""
            marker = f"{domain}|{career_url}"
            if not domain or marker in seen:
                continue
            seen.add(marker)
            inputs.append(UploadDomainInput(domain=domain, career_url=career_url or None))
        return inputs

    def _optional_header_index(self, headers: list[str], choices: list[str]) -> int | None:
        for choice in choices:
            if choice in headers:
                return headers.index(choice)
        return None

    def _optional_header_name(self, headers: dict[str, str], choices: list[str]) -> str | None:
        for choice in choices:
            if choice in headers:
                return headers[choice]
        return None
